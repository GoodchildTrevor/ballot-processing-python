import pandas as pd
from typing import List, Dict, Any, Tuple, Optional
import os
from copy import deepcopy


def delete_non_breaking_spaces(text: Any) -> str:
    """Replace non-breaking spaces and strip."""
    if isinstance(text, str):
        return text.replace("\u00A0", " ").strip()
    return str(text)


def postfix(score: int) -> str:
    """Return correct Russian postfix for score."""
    if score == 1:
        return f"{score} балл"
    elif 2 <= score <= 4:
        return f"{score} балла"
    else:
        return f"{score} баллов"


def process_coincidences(nominations_dict: Dict[str, Dict]) -> List[Tuple]:
    """
    Find people in both main and supporting categories.
    Returns: [(name, total_score, main_score, support_score, mentions, category)]
    """
    main_to_support = {"actor": "actor2", "actress": "actress2"}
    coincidences = []

    for main_cat, support_cat in main_to_support.items():
        main_data = nominations_dict.get(main_cat, {})
        support_data = nominations_dict.get(support_cat, {})

        for name, main_info in main_data.items():
            if name in support_data:
                total = main_info["score"] + support_data[name]["score"]
                mentions = main_info["mentions"] + support_data[name]["mentions"]
                coincidences.append((
                    name,
                    total,
                    main_info["score"],
                    support_data[name]["score"],
                    mentions,
                    main_cat
                ))
    return coincidences


def process_data(
    df_original: pd.DataFrame,
    nominees_df: pd.DataFrame,
    selected_nominations: List[str]
) -> Tuple[Dict, Dict]:
    """
    Process voting data.
    Returns: (movies_dict, nominations_dict)
    """
    # Replace NaN with "xxx"
    df = df_original.fillna("xxx")
    
    # Voter columns: skip first column ("Ваш ник...")
    voter_cols = df.columns[1:]
    df_transposed = df.set_index(df.columns[0]).T  # now rows = voters, cols = rankings

    # Top 10 films (rows 0–9)
    df_movies = df_transposed.iloc[:, :10]
    # Other rows (nominations)
    df_other = df_transposed.iloc[:, 10:]

    # === Movies scoring ===
    movies_dict = {}
    for voter, row in df_movies.iterrows():
        for pos in range(10):
            movie = str(row.iloc[pos])
            score = 11 - (pos + 1)
            if movie not in movies_dict:
                movies_dict[movie] = {"score": 0, "mentions": []}
            movies_dict[movie]["score"] += score
            movies_dict[movie]["mentions"].append((voter, postfix(score)))

    # === Nominations scoring ===
    nominations_dict = {nom: {} for nom in selected_nominations}

    # Ensure nominees_df has same number of columns as selected_nominations
    nominees_list = []
    for i, nom in enumerate(selected_nominations):
        if i < len(nominees_df.columns):
            col = nominees_df.iloc[:, i].dropna()
            nominees_list.append([delete_non_breaking_spaces(x) for x in col])
        else:
            nominees_list.append([])

    # For each nomination category
    for nom_idx, nom in enumerate(selected_nominations):
        nominees_in_category = nominees_list[nom_idx]
        if not nominees_in_category:
            continue

        # This category corresponds to row `nom_idx` in df_other
        if nom_idx >= len(df_other.columns):
            continue

        for nominee in nominees_in_category:
            for voter, row in df_other.iterrows():
                value = str(row.iloc[nom_idx])
                if nominee in value:
                    if nominee not in nominations_dict[nom]:
                        nominations_dict[nom][nominee] = {"score": 0, "mentions": []}
                    nominations_dict[nom][nominee]["score"] += 1
                    nominations_dict[nom][nominee]["mentions"].append(voter)

    return movies_dict, nominations_dict


def write_results_to_excel(
    input_path: str,
    output_path: str,
    all_results: List[Tuple[Optional[Dict], Dict]],
    sheet_names: List[str],
    selected_nominations: List[str]
) -> None:
    """Write all results to Excel file."""
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Copy input file to output
    import shutil
    shutil.copy2(input_path, output_path)

    # Load workbook
    wb = load_workbook(output_path)

    # Remove existing result sheets (if any)
    for sheet in sheet_names:
        if sheet in wb.sheetnames:
            del wb[sheet]

    # Process each sheet
    for sheet_name, (movies_dict, nominations_dict) in zip(sheet_names, all_results):
        ws = wb.create_sheet(title=sheet_name)

        if sheet_name == "совпадения":
            # Write coincidences
            coincidences = process_coincidences(nominations_dict)
            headers = ["Имя", "Баллы (итого)", "Первый план", "Второй план", "Упоминания"]
            ws.append(headers)
            for name, total, main, support, mentions, _ in coincidences:
                ws.append([
                    name,
                    total,
                    main,
                    support,
                    ", ".join(mentions)
                ])
        else:
            # Best movies
            if movies_dict is not None:
                sorted_movies = sorted(
                    [(k, v) for k, v in movies_dict.items() if k != "xxx"],
                    key=lambda x: x[1]["score"],
                    reverse=True
                )
                for i, (name, info) in enumerate(sorted_movies, start=1):
                    mentions_str = ", ".join(f"{v} ({m})" for v, m in info["mentions"])
                    ws.cell(row=i+1, column=1, value=name)
                    ws.cell(row=i+1, column=2, value=info["score"])
                    ws.cell(row=i+1, column=3, value=mentions_str)

            # Headers for nominations
            nomination_plus = ["movie"] + selected_nominations
            for z, nom in enumerate(nomination_plus):
                ws.cell(row=1, column=z*3+1, value=nom)
                ws.cell(row=1, column=z*3+2, value=f"points_{nom}")
                ws.cell(row=1, column=z*3+3, value=f"mentions_by_{nom}")

            # Nomination results
            for col_idx, nom in enumerate(selected_nominations):
                if nom not in nominations_dict:
                    continue
                data = nominations_dict[nom]
                sorted_data = sorted(
                    data.items(),
                    key=lambda x: x[1]["score"],
                    reverse=True
                )
                for line_idx, (name, info) in enumerate(sorted_data):
                    mentions_str = ", ".join(info["mentions"])
                    ws.cell(row=line_idx+2, column=4 + col_idx*3, value=name)
                    ws.cell(row=line_idx+2, column=4 + col_idx*3 + 1, value=info["score"])
                    ws.cell(row=line_idx+2, column=4 + col_idx*3 + 2, value=mentions_str)

    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)


def run_voting(input_path: str, output_path: str, selected_nominations: List[str]) -> None:
    """
    Full pipeline: read, process, write.
    """
    # Read data
    df_original = pd.read_excel(input_path, sheet_name="номинанты")
    nominees_df = pd.read_excel(input_path, sheet_name="списки")

    # Auto-generate slices
    n_votes = len(df_original.columns) - 1  # minus "Ваш ник..."
    max_slice = (n_votes // 10) * 10
    slices = list(range(10, max_slice + 1, 10)) if max_slice >= 10 else []

    # Full processing
    final_movies, final_nominations = process_data(df_original, nominees_df, selected_nominations)

    all_results = [(final_movies, final_nominations)]

    # Process slices
    for s in slices:
        # Take first `s` voters
        voter_cols = ["Ваш ник на Форуме Кинопоиска:"] + list(df_original.columns[1:s+1])
        subset_df = df_original[voter_cols]
        movies, nominations = process_data(subset_df, nominees_df, selected_nominations)
        all_results.append((movies, nominations))

    # Add coincidences (only needs nominations dict)
    all_results.append((None, final_nominations))

    # Sheet names
    winner_sheet_names = ["победители"]
    if slices:
        winner_sheet_names += [f"победители {s}" for s in slices]
    sheet_names = winner_sheet_names + ["совпадения"]

    # Write to Excel
    write_results_to_excel(input_path, output_path, all_results, sheet_names, selected_nominations)