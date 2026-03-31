import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any


def delete_non_breaking_spaces(text: str) -> str:
    """
    Remove non-breaking spaces and extra whitespace from text.

    :param text: Input text or value to clean
    :type text: str
    :return: Cleaned text with regular spaces only
    :rtype: str

    :Example:

    >>> delete_non_breaking_spaces("Hello\\u00A0World")
    'Hello World'
    >>> delete_non_breaking_spaces(None)
    ''
    """
    if pd.isna(text):
        return ""
    return str(text).replace("\u00A0", " ").replace("  ", " ").strip()


def postfix(score: int) -> str:
    """
    Generate Russian grammatical form for score with 'балл/балла/баллов'.

    :param score: Integer score value
    :type score: int
    :return: Formatted string with proper Russian plural form
    :rtype: str

    :Example:

    >>> postfix(1)
    '1 балл'
    >>> postfix(3)
    '3 балла'
    >>> postfix(10)
    '10 баллов'
    """
    if score == 1:
        return f"{score} балл"
    elif 2 <= score <= 4:
        return f"{score} балла"
    else:
        return f"{score} баллов"


def process_data(
    df: pd.DataFrame,
    nominees_df: pd.DataFrame,
    nominations: list[str]
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, dict[str, Any]]]]:
    """
    Process voting data and calculate scores for movies and nominations.

    :param df: DataFrame containing voter names in first column and votes in remaining columns.
               First 10 vote columns are movie rankings (1st-10th place).
               Subsequent columns correspond to nomination categories.
    :type df: pd.DataFrame
    :param nominees_df: DataFrame with nominee names, one column per nomination category
    :type nominees_df: pd.DataFrame
    :param nominations: List of nomination category names (must match nominees_df columns)
    :type nominations: List[str]
    :return: Tuple of (movies_dict, nominations_dict) where:

             - movies_dict maps movie names to {"score": int, "mentions": List[str]}
             - nominations_dict maps category -> nominee -> {"score": int, "mentions": List[str]}
    :rtype: Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Dict[str, Any]]]]

    :Example:

    >>> df = pd.DataFrame({
    ...     'voter': ['Alice', 'Bob'],
    ...     'movie1': ['Film A', 'Film B'],
    ...     'actor': ['Actor X', 'Actor Y']
    ... })
    >>> nominees = pd.DataFrame({'actor': ['Actor X', 'Actor Y']})
    >>> movies, noms = process_data(df, nominees, ['actor'])
    """
    df = df.fillna("xxx")

    voter_names: list[str] = df.iloc[:, 0].astype(str).tolist()

    df_votes: pd.DataFrame = df.iloc[:, 1:]

    movies_dict: dict[str, dict[str, Any]] = {}
    df_movies: pd.DataFrame = df_votes.iloc[:, 0:10]

    for voter_idx, row in enumerate(df_movies.values):
        voter_name: str = voter_names[voter_idx]
        for pos in range(10):
            movie_key: str = str(row[pos])
            score: int = 10 - pos

            if movie_key not in movies_dict:
                movies_dict[movie_key] = {"score": 0, "mentions": []}

            movies_dict[movie_key]["score"] += score
            movies_dict[movie_key]["mentions"].append(f"{voter_name} ({postfix(score)})")

    nominations_dict: dict[str, dict[str, dict[str, Any]]] = {nom: {} for nom in nominations}
    df_other: pd.DataFrame = df_votes.iloc[:, 10:]

    for i, nom_cat in enumerate(nominations):
        if i < len(nominees_df.columns):
            current_nominees: pd.Series = nominees_df.iloc[:, i].dropna().unique()

            for nominee in current_nominees:
                clean_nominee: str = delete_non_breaking_spaces(nominee)
                if not clean_nominee:
                    continue

                for voter_idx, vote_val in enumerate(df_other.iloc[:, i].values):
                    if clean_nominee in str(vote_val):
                        if clean_nominee not in nominations_dict[nom_cat]:
                            nominations_dict[nom_cat][clean_nominee] = {"score": 0, "mentions": []}

                        nominations_dict[nom_cat][clean_nominee]["score"] += 1
                        nominations_dict[nom_cat][clean_nominee]["mentions"].append(voter_names[voter_idx])

    return movies_dict, nominations_dict


def process_coincidences(
    nominations_dict: dict[str, dict[str, dict[str, Any]]]
) -> list[tuple[str, int, int, int, str]]:
    """
    Identify actors/actresses who appear in both main and supporting categories.

    :param nominations_dict: Dictionary mapping nomination categories to nominees and their scores
    :type nominations_dict: Dict[str, Dict[str, Dict[str, Any]]]
    :return: List of tuples (name, total_score, main_score, supporting_score, combined_mentions)
    :rtype: List[Tuple[str, int, int, int, str]]

    :note: Checks 'actor' vs 'actor2' and 'actress' vs 'actress2' categories
    """
    main_cats: list[str] = ["actor", "actress"]
    supp_cats: list[str] = ["actor2", "actress2"]
    coincidences: list[tuple[str, int, int, int, str]] = []

    for m, s in zip(main_cats, supp_cats):
        data_main: dict[str, dict[str, Any]] = nominations_dict.get(m, {})
        data_supp: dict[str, dict[str, Any]] = nominations_dict.get(s, {})

        for name, info in data_main.items():
            if name in data_supp:
                total_score: int = info["score"] + data_supp[name]["score"]
                combined_mentions: list[str] = info["mentions"] + data_supp[name]["mentions"]
                coincidences.append((
                    name, total_score, info["score"], data_supp[name]["score"],
                    ", ".join(combined_mentions)
                ))
    return coincidences


def run_voting(file_path: str, nominations: list[str]) -> None:
    """
    Run complete voting analysis and update Excel file with results.

    Reads voting data from Excel file with two sheets:
    - 'номинанты': voter names and votes
    - 'списки': nominee lists for each category

    Creates/updates result sheets:
    - 'победители': final results with all votes
    - 'победители N': intermediate results after N votes (every 10 votes)
    - 'совпадения': actors appearing in both main and supporting categories

    :param file_path: Path to Excel file containing voting data
    :type file_path: str
    :param nominations: List of nomination category names
    :type nominations: List[str]
    :return: None (modifies Excel file in place)
    :rtype: None
    :raises FileNotFoundError: If file_path does not exist
    :raises KeyError: If required sheets are missing

    :Example:

    >>> run_voting('votes.xlsx', ['actor', 'actress', 'director'])
    """
    df_original: pd.DataFrame = pd.read_excel(file_path, sheet_name="номинанты")
    nominees_raw: pd.DataFrame = pd.read_excel(file_path, sheet_name="списки")

    nominees_df: pd.DataFrame = nominees_raw.copy()
    nominees_df.columns = nominations[:len(nominees_df.columns)]

    n_votes: int = len(df_original)
    max_slice: int = (n_votes // 10) * 10
    slices: list[int] = list(range(10, max_slice + 1, 10)) if max_slice >= 10 else []

    final_movies, final_noms = process_data(df_original, nominees_df, nominations)

    results_map: list[tuple[str, tuple[dict, dict]]] = [("победители", (final_movies, final_noms))]
    for s in slices:
        subset: pd.DataFrame = df_original.iloc[:s, :]
        results_map.append((f"победители {s}", process_data(subset, nominees_df, nominations)))

    wb: Workbook = openpyxl.load_workbook(file_path)

    for sheet_name, (m_dict, n_dict) in results_map:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws: Worksheet = wb.create_sheet(sheet_name)

        sorted_movies: list[str] = sorted(
            [k for k in m_dict if k != "xxx"],
            key=lambda x: m_dict[x]["score"],
            reverse=True
        )

        headers: list[str] = ["movie"] + nominations
        for z, h in enumerate(headers):
            col: int = z * 3 + 1
            ws.cell(1, col, h)
            ws.cell(1, col + 1, f"points_{h}")
            ws.cell(1, col + 2, f"mentions_by_{h}")

        for row_idx, movie in enumerate(sorted_movies, start=2):
            ws.cell(row_idx, 1, movie)
            ws.cell(row_idx, 2, m_dict[movie]["score"])
            ws.cell(row_idx, 3, ", ".join(m_dict[movie]["mentions"]))

        for col_idx, nom in enumerate(nominations):
            start_col: int = 4 + (col_idx * 3)
            sorted_noms: list[tuple[str, dict[str, Any]]] = sorted(
                n_dict[nom].items(),
                key=lambda x: x[1]["score"],
                reverse=True
            )

            for row_offset, (name, info) in enumerate(sorted_noms, start=2):
                ws.cell(row_offset, start_col, name)
                ws.cell(row_offset, start_col + 1, info["score"])
                ws.cell(row_offset, start_col + 2, ", ".join(info["mentions"]))

    sn_coinc: str = "совпадения"
    if sn_coinc in wb.sheetnames:
        del wb[sn_coinc]
    ws_c: Worksheet = wb.create_sheet(sn_coinc)
    ws_c.append(["Имя", "Баллы (итого)", "Первый план", "Второй план", "Упоминания"])

    coincidences: list[tuple[str, int, int, int, str]] = process_coincidences(final_noms)
    for row in coincidences:
        ws_c.append(row)

    wb.save(file_path)
